#!/usr/bin/env python
# coding: utf-8

# # LIST

# #### Q1. Write a Python program which accepts a sequence of comma-separated numbers from user and generate a list and a tuple with those numbers.
# 
# Sample data : 3, 5, 7, 23
# Output :
# List : ['3', ' 5', ' 7', ' 23']
# Tuple : ('3', ' 5', ' 7', ' 23')

# In[1]:


numbers=input('Comma Seperated Numbers :').split(',')
print('List :',numbers)
print('Tuple :',tuple(numbers))


# #### Q2. Write a Python program to display the first and last colors from the following list.
# 
# color_list = ["Red","Green","White" ,"Black"]

# In[2]:


color_list=['Red','Green','White','Black']
print(color_list[0])
print(color_list[-1])


# #### Q3. Write a Python program to print the even numbers from a given list.
# 
# Sample List : [1, 2, 3, 4, 5, 6, 7, 8, 9]
# Expected Result : [2, 4, 6, 8]

# In[6]:


Sample_List=[1,2,3,4,5,6,7,8,9]
print([x for x in Sample_List if x%2==0])


# # MODULE

# #### Q1. Write a Python program to calculate number of days between two dates. Hint: use Datetime package/module.
# 
# Sample dates : (2014, 7, 2), (2014, 7, 11)
# Expected output : 9 days

# In[12]:


from datetime import datetime

str_d1='2014/07/02'
str_d2='2014/07/11'

#converting string format into date format
d1=datetime.strptime(str_d1,"%Y/%m/%d")
d2=datetime.strptime(str_d2,"%Y/%m/%d")

#difference between dates
delta = d2-d1
print(f'Difference is {delta.days} days')


# ## FUNCTIONS

# #### Q1. Write a Python program to get the volume of a sphere with radius 6.
# 
# 

# In[15]:


import math

def volume_of_sphere_with_radius(radius):
    v=4/3*math.pi*radius**3
    return v
volume_of_sphere_with_radius(6)


# #### Q2. Write a Python program to calculate the sum of three given numbers, if the values are equal then return three times of their sum hint: write User defined functions

# In[20]:


def sum_of_three(a, b, c):
    a = int(a)
    b = int(b)
    c = int(c)
    if a == b and b== c and a == c:
        return (a+b+c)*3
    else:
        return a+b+c

sum_of_three(4,4,4)


# In[21]:


sum_of_three(2,5,3)


# #### Q3. Write a Python program to count the number 4 in a given list.
# 
# List = [1,4,6,8,4,9,4]

# In[45]:


List = [1,4,6,8,4,9,4]

def count_four(lists):
    four = 0
    for i in lists:
        if i == 4 :
            four= four + i/4
    print("Count of 4 in the list :",int(four))      


# In[46]:


count_four(List)


# #### Q4. Write a Python program to print all even numbers from a given numbers list in the same order and stop the printing if any numbers that come after 237 in the sequence. Go to the editorSample numbers list :
# 
# 399, 162, 758, 219, 918, 237, 412, 566, 826, 248, 866, 950, 626, 949, 687, 217,
# 
# 815, 67, 104, 58, 512, 24, 892, 894, 767, 553, 81, 379, 843, 831, 445, 742, 717,
# 
# 958,743, 527]

# In[58]:


numbers = [399, 162, 758, 219, 918, 237, 412, 566, 826, 248, 866, 950, 626, 949, 687, 217,

815, 67, 104, 58, 512, 24, 892, 894, 767, 553, 81, 379, 843, 831, 445, 742, 717,

958,743, 527] 


def even_numbers(numbers):
    for x in numbers:
        if x % 2 == 0:
             print(x)
        elif x == 237:
            break

even_numbers(numbers)


# #### Q5. Write a Python program to find those numbers which are divisible by 7 and multiple of 5, between 1500 and 2700 (both included)
# 
# 

# In[61]:


print([ i for i in range(1500, 2701) if i % 7 == 0 and i % 5 == 0] )


# #### Q6. Write a Python program that prints all the numbers from 0 to 6 except 3 and 6.
# 
# Note : Use 'continue' statement.
# 
# Expected Output : 0 1 2 4 5

# In[74]:


for i in range(7):
    if i == 3 or i == 6:
        continue
    print(i, end = ' ')
 


# #### Q7. Write a Python program to get the Fibonacci series between 0 to 50.
# 
# Note : The Fibonacci Sequence is the series of numbers :
# 
# 0, 1, 1, 2, 3, 5, 8, 13, 21, ....
# 
# Every next number is found by adding up the two numbers before it.
# 
# Expected Output : 1 1 2 3 5 8 13 21 34

# In[78]:


# Program to display the Fibonacci sequence up to n-th term

nterms = int(input("How many terms? "))

# first two terms
n1, n2 = 0, 1
count = 0

# check if the number of terms is valid
if nterms <= 0:
    print("Please enter a positive integer")
# if there is only one term, return n1
elif nterms == 1:
    print("Fibonacci sequence upto",nterms,":")
    print(n1)
# generate fibonacci sequence
else:
    print("Fibonacci sequence:")
    while count < nterms:
        print(n1)
        nth = n1 + n2
       # update values
        n1 = n2
        n2 = nth
        count += 1


# #### Q8. Write a Python program to get the Fibonacci series between 0 to 50.
# 
# Note : The Fibonacci Sequence is the series of numbers :
# 
# 0, 1, 1, 2, 3, 5, 8, 13, 21, ....
# 
# Every next number is found by adding up the two numbers before it.
# 
# Expected Output : 1 1 2 3 5 8 13 21 34

# In[79]:


nterms = [0,1,1,2,3,5,8,13,21]

n1, n2 = 0, 1
count = 0

print("Fibonacci sequence:")
while count < len(nterms):
    print(n2, end = ' ')
    nth = n1 + n2
    n1 = n2
    n2 = nth
    count += 1


# #### Q9. Write a Python function that takes a list and returns a new list with unique elements of the first list.
# 
# Sample List : [1,2,3,3,3,3,4,5]
# 
# Unique List : [1, 2, 3, 4, 5]
# 
# 

# In[80]:


def unique_(lists):
    a = []
    for i in lists:
        if i not in a:
            a.append(i)
    return a

sample_list = [1,2,3,3,3,3,4,5]

unique_(sample_list)


# ## STRINGS

# #### Q1. Write a Python program to concatenate all elements in a list into a string and return it.
# 
# 

# In[83]:


list=['My','Name','Is','Vaibhavi']
string=" ".join([str(i) for i in list])
print(string)


# ## DICTIONARY

# #### Q1. Write a Python script to concatenate following dictionaries to create a new one.
# 
# Sample Dictionary :
# 
# dic1={1:10, 2:20}
# 
# dic2={3:30, 4:40}
# 
# dic3={5:50,6:60}
# 
# Expected Result : {1: 10, 2: 20, 3: 30, 4: 40, 5: 50, 6: 60}

# In[88]:



dic1={1:10, 2:20}

dic2={3:30, 4:40}

dic3={5:50,6:60}

dic1.update(dic2)
dic1
dic1.update(dic3)
dic1


# ## SERIES

# #### Q1. Write a Python program to add, subtract, multiple and divide two Pandas Series.
# 
# Sample Series: [2, 4, 6, 8, 10], [1, 3, 5, 7, 9]

# In[90]:


def arithmetic_func(l1, l2, function):
    l3 = []
    if len(l1) != len(l2):
        print("provide list with equal length")
    elif function == 'add':
        for i in range(len(l1)):
            l3.append(l1[i] + l2[i])
        return l3
    elif function == 'subtract':
        for i in range(len(l1)):
            l3.append(l1[i] - l2[i])
        return l3
    elif function == 'multiply':
        for i in range(len(l1)):
            l3.append(l1[i] * l2[i])
        return l3
    elif function == 'divide':
        for i in range(len(l1)):
            l3.append(round(l1[i] / l2[i],2))
        return l3
    


# In[99]:


l1=[2,4,6,8,10]
l2=[1,3,5,7,9]
arithmetic_func(l1, l2, 'add')


# In[96]:


arithmetic_func(l1, l2, 'subtract')


# In[97]:


arithmetic_func(l1, l2, 'multiply')


# In[98]:


arithmetic_func(l1, l2, 'divide')


# ## DATA FRAME

# #### Q1. Write a Pandas program to select the specified columns and rows from a given data frame. Go to the editorSample Python dictionary data and list labels:
# 
# Select 'name' and 'score' columns in rows 1, 3, 5, 6 from the following data frame.
# 
# exam_data = {'name': ['Anastasia', 'Dima', 'Katherine', 'James', 'Emily', 'Michael', 'Matthew', 'Laura', 'Kevin', 'Jonas'],
# 
# score': [12.5, 9, 16.5, np.nan, 9, 20, 14.5, np.nan, 8, 19],
# 
# attempts': [1, 3, 2, 3, 2, 3, 1, 1, 2, 1],
# 
# qualify': ['yes', 'no', 'yes', 'no', 'no', 'yes', 'yes', 'no', 'no', 'yes']}
# 
# labels = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j']
# 
# Expected Output:
# 
# Select specific columns and rows:
# 
# name score
# 
# b Dima 9.0
# 
# d James NaN
# 
# f Michael 20.0
# 
# g Matthew 14.5

# In[7]:


import numpy as np
import pandas as pd

exam_data ={'name': ['Anastasia', 'Dima', 'Katherine', 'James', 'Emily', 'Michael', 'Matthew', 'Laura', 'Kevin', 'Jonas'],

'score':[12.5, 9, 16.5, np.nan, 9, 20, 14.5, np.nan, 8, 19],

'attempts' :[1, 3, 2, 3, 2, 3, 1, 1, 2, 1],

'qualify':['yes', 'no', 'yes', 'no', 'no', 'yes', 'yes', 'no', 'no', 'yes']}

labels=  ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j']

df = pd.DataFrame(exam_data, index=labels)
df


# In[8]:


df.iloc[[1,3,5,6],[0,1]]


# #### Q2. Use Crime dataset from LMS
# 
# I) find the aggregations like all moments of business decisions for all columns,value counts.
# 
# II) do the plottings like plottings like histogram, boxplot, scatterplot, barplot, piechart,dot chart.

# #### I) find the aggregations like all moments of business decisions for all columns,value counts.
# 

# In[11]:


import pandas as pd

crime_dataset = pd.read_csv(r"C:\Users\Vaibhavi\OneDrive\Desktop\crime_data.csv")
print(crime_dataset.describe())


# #### II) do the plottings like plottings like histogram, boxplot, scatterplot, barplot, piechart,dot chart.

# In[29]:


import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt

plt.hist(crime_dataset, density = True, 
         histtype ='bar')

plt.legend(prop ={'size': 10})
  
plt.title('matplotlib.pyplot.hist() function Example\n\n',
          fontweight ="bold")
  
plt.show()


# #### Q3. use mtcars dataset from LMS
# 
# A) delete/ drop rows-10 to 15 of all columns
# 
# B)drop the VOL column
# 
# C)write the forloop to get value_counts of all cloumns

# In[24]:


mtc=pd.read_csv(r"C:\Users\Vaibhavi\OneDrive\Desktop\mtcars.csv")
mtc


# #### A) delete/ drop rows-10 to 15 of all columns

# In[25]:


mtc.drop(index = list(range(10,16)))


# #### B)drop the VOL column
# 

# In[28]:


mtc.drop(columns = 'VOL')


# #### C)write the forloop to get value_counts of all cloumns

# In[30]:


for column in mtc.columns:
    print("Column Name: ", column)
    print(mtc[column].value_counts())


# #### Q4. Use Bank Dataset from LMS
# 
# A)change all the categorical columns into numerical by creating Dummies and using label encoder.
# 
# B) rename all the column names DF
# 
# C) Rename only one specific column in DF

# #### A)change all the categorical columns into numerical by creating Dummies and using label encoder.
# 

# In[3]:


import pandas as pd
bank_data=pd.read_csv(r"C:\Users\Vaibhavi\OneDrive\Desktop\bank-full.csv",delimiter=';')
bank_data


# In[45]:


#by using dummies

df= bank_data.copy()
df
# One-Hot Encoding of categrical variables
df=pd.get_dummies(bank_data,columns=['job', 'marital', 'education', 'contact', 'month', 'poutcome'])
df
pd.set_option("display.max.columns", None)
df


# In[14]:


#by using label - encoder 
import pandas as pd
import numpy as np
from sklearn.preprocessing import LabelEncoder

#creating initial dataframe
b=('job', 'marital', 'education', 'contact', 'month', 'poutcome')
b
bank_df=pd.DataFrame(b,columns=['Bank_columns'])

# creating instance of labelencoder
labelencoder = LabelEncoder()

# Assigning numerical values and storing in another column
bank_df['Bank_categorical'] = labelencoder.fit_transform(bank_df['Bank_columns'])
bank_df


# #### B) rename all the column names DF
# 
# 

# In[8]:


print('Old column names: ',bank_data.columns)


# In[12]:


bank_data.rename({'age':'AGE','job':'JOB',                  'marital':'MARITAL',                  'education':'EDUCATION',                  'default':'DEFAULT',                  'balance':'BALANCE',                  'housing':'HOUSING',                  'loan':'LOAN',                  'contact':'CONTACT',                  'day':'DAY',                  'month':'MONTH',                  'duration':'DURATION',                  'campaign':'CAMPAIGN',                  'pdays':'PDAYS',                  'previous':'PREVIOUS',                  'poutcome':'POUTCOME',                  'y':'Y'} , axis =1 , inplace = True)
bank_data


# #### C) Rename only one specific column in DF
# 
# 

# In[26]:


bank_data.rename({'EDUCATION':'ED'}, axis =1 , inplace = True)
bank_data.columns


# #### Q5. After doing all the changes in bank data(Q19). save the file in your directory in Csv Format.
# 
# 

# In[31]:


bank_data.to_csv("bank_data.csv")


# ## BASIC PROGRAMS

# #### Q1. Write Python Programs to use various operators in Python
# 
# 

# In[55]:


def various_operators(list1,list2,function):
    list3=[]
    if len(list1)!= len(list2):
        print ("Provide both the lists with equal length")
    elif function == 'add':
        for i in range (len(list1)):
            list3.append(list1[i]+list2[i])
        return list3
    elif function == 'subtract':
        for i in range (len(list1)):
            list3.append(list1[i]-list2[i])
        return list3
    elif function == 'multiply':
        for i in range(len(list1)):
            list3.append(list1[i]*list2[i])
        return list3
    elif function == 'divide':
        for i in range(len(list1)):
            list3.append(list1[i]/list2[i])
        return list3
    elif function == 'modulo' :
        for i in range(len(list1)):
            list3.append(list1[i]%list2[i])
        return list3
    else:
        print('Please provide a valid arithmetic operator like add, subtract, multiply,divide , modulo')

        
        
        
        


# In[46]:


list1=[2,2,2]
list2=[3,3,3]
various_operators(list1,list2,'add')


# In[58]:


list1=[2,8,10]
list2=[1,3,7]
various_operators(list1,list2,'subtract')


# In[53]:


list1=[2,1,2]
list2=[3,3,4]
various_operators(list1,list2,'multiply')


# In[57]:


list1=[4,24,6]
list2=[2,4,3]
various_operators(list1,list2,'divide')


# #### Q2. Create list of elements and slice and dice it
# 
# 

# In[63]:


List = ['apple','banana','mango','chiku','pear','anjir','papaya']
List


# In[64]:


List[:7]


# In[65]:


List[3:5]


# In[67]:


List[:-2]


# In[68]:


List[2:7]


# #### Q3. Using while loop accept numbers until sum of numbers is less than 100
# 
# 

# In[7]:


i = 0
num = 0
while i < 100:
    num += 1
    Sum = num
    i = Sum + num

print(num, Sum, i)


# #### Q4. Write a python program Read & write Excel files 
# 
# 

# In[3]:


pip install openpyxl


# In[11]:


from openpyxl import Workbook
import time

wb =Workbook()
sheet = wb.active

sheet['A1'] = 'Name'
sheet['A2'] = 'Jack'
sheet['A3'] = 29

wb.save("workbook.xlsx")


# In[14]:


#import libraries
from openpyxl import load_workbook

w = load_workbook("workbook.xlsx")
sheet = w.active


# #### 5. Write a python program to scrape reviews from a commercial web site
# 
# 

# In[ ]:





# #### Q6. Create a 3x3 matrix with values ranging from 2 to 10 using numpy
# 
# 

# In[8]:


import numpy as np
x =  np.arange(2, 11).reshape(3,3)
print(x)


# #### Q7. Write a Python program to convert a list of numeric value into a one-dimensional NumPy array
# 
# 

# In[11]:


import numpy as np

list = [24, 32.5, 100, 12.36]
print("Original List:",list)
array = np.array(list)
print("One-dimensional NumPy array: ",array)


# #### Q8.Write a Python program to create a null vector of size 10 and update sixth value to 11.
# 
# 

# In[15]:


import numpy as np

x=np.zeros(10)
print(x)

x[6]=11
print(x)

